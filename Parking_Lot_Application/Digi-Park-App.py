#Preston Ott, OYSMS, Digi-Park, 4/19/2023
#This is a update to the parking app from version 6 to version 7, adds a new file structure to work on any machine


#Preston Ott, OYSMS, Digi-Park, Update to Parking App, 4/14/2023
#This is a update to the parking app version V8,

#Importing the Kivy Classes We'll need to Build the App
import kivy
import os
import math
from kivy.app import App
from kivy.uix.widget import Widget
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.core.window import Window
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.image import Image
from kivy.uix.textinput import TextInput
from kivy.uix.scrollview import ScrollView
import openpyxl as xl
#3/20/2023 socket for client functionality
import socket

#Importing PILLOW for Image Processing for App
from PIL import Image as im



#Here the park_app will have 2 arrays (address_array) and (lotimage_lists) these will contain the addresses digipark is implemented in
#lotimage_lists will contain the parking lot empty map images we will display to the user. These will need to be updated by the server

#Making a (root or reference( directory that all the files and data of our app stems from
ref = os.getcwd()

#4/18/2023 Update Preston-Ott
#1.) I need to add in the arrays for the addresses via the directory Parking_Application/address_list
f1 = open("address_list.txt",'r')
address_list = []
#Reading the text file line by line
for line in (f1):
    address_list.append(line.strip())

#2.) I need to add in the empty map images via the directory Parking_Aplication/Empty_Maps/Emptymap_name
images = os.listdir("Empty_Maps")
os.chdir(r"Empty_Maps")
print(images)
image_list = []
for i in range (len(images)):
    image_list.append(im.open(images[i]))
os.chdir(ref)

#3.) I need to add in the marker images via the directory Parking_Application/Marker_Images
markers = os.listdir("Marker_Images")
os.chdir("Marker_Images")
marker_list = []
for i in range(len(markers)):
    marker_list.append(im.open(markers[i]))
os.chdir(ref)

#4.) I need to add in the position arrays for each parking lot map via the directorym Parking/Application/Lot_Images_Pos
#These will be a bunch of excel files so I need to read them one by one, and convert them into a pos_list[]
posmaps = os.listdir("Lot_Image_Pos")
os.chdir(r"Lot_Image_Pos")
pos_list = []
for i in range(len(posmaps)):
    #Getting excel file, and worksheet
    wb = xl.load_workbook(posmaps[i])
    sheet = wb.worksheets[0]
    #Getting size of workbook
    size = sheet.max_row
    #Creating temp pos_list
    pos_array = []
    for j in range(size):
        #Adding the pixels pairs
        pos_array.append((sheet.cell(row = j+1,column = 1).value, sheet.cell(row = j+1, column = 2).value))
    #Adding the pos_list to our cell
    pos_list.append(pos_array)

#Returning t reference directory
os.chdir(ref)
    
#5.) Fetching  the list of ip addresses the user will use to connect to the correct server for a specific parking lot, 4/20/2023
ip_addr_list = []
f1 = open(r"C:\Users\large\OneDrive\Documents\Desktop\Design Project\Parking_App_Versions\Parking_Application\server_ip_list.txt")
for line in (f1):
    ip_addr_list.append(line.strip())
#Returning to reference directory
os.chdir(ref)

#All functions that used address_list[], images_array, marker_list[0], marker_list[1], pos_list[], will need to be modified to work with our map

#Here we define a function that will take in a (parkinglot image) and (paste 1) marker at a specific location and return the modified image
def mark_spot(im1,red,green,state,pos):
    
    #Here we define a new image the same size as im1, and paste the map im1
    im_new = im.new("RGB",im1.size)
    im_new.paste(im1,(0,0))
    
    #Now we paste the red or the green marker at a pos to mark a spot
    if(state == 0):
        im_new.paste(red,pos)
    else:
        im_new.paste(green,pos)
        
    #Returning the updated image
    return im_new

#Here define a function that will draw the map using (spot_array), (pos_list) that tell us the (state of each spot),(pixel pos of each spot)
def draw_map(im1,red,green,spot,pos):
    
    #Here we create a new image to store the map image
    new_im = im.new("RGB",im1.size)
    #Pasting empty map into new image
    new_im.paste(im1,(0,0))
    
    #We enter a loop to mark each spot in sequential order
    print(len(spot))
    for i in range(len(spot)):
        #Updating the new_image
        new_im = mark_spot(new_im,red,green,spot[i],pos[i])
    
    #Returning the marked parking map image
    return new_im



#Here we make a sub-class of the App class for Parking App (Added address_array and image_list as parameters)
class park_app(App):
    
    #Rewriting the Build Function Which Creates the Root Widget
    def build(self):
        
        #Creating the Main Menu Widget Object
        main1 = mm()
        
        #Returning the Main Menu Widget to be runned
        return main1
        
        

#Here we create a sub-class of the widget class for the main-menu of app
class mm(GridLayout):

    
    #Initializing the widget object
    def __init__(self,**kwargs):
        
        #Calling the super classes constructor
        super(mm,self).__init__(**kwargs,rows = 4,cols = 2,orientation = 'tb-lr')
        
        self.count = 0
        
        #Here we add a label for the app
        app_label = Label(text = "DigiPark", font_size = 75)
        self.add_widget(app_label)
        
        #Here we create 3 button widgets to the mm widget
        findlot = Button(text = "Find Lot",font_size = 75)
        profile = Button(text = "Profile",font_size = 75)
        exit1 = Button(text = "Exit", font_size = 75)
        
        #Adding the 3 Buttons to the main_menu widget
        self.add_widget(findlot)
        self.add_widget(profile)
        self.add_widget(exit1)
        
        
        #Binding the buttons to certain functions
        exit1.bind(on_press = self.end)
        findlot.bind(on_press = self.open_flot)
        profile.bind(on_press = self.open_profile)
        
    #Making a function to bind to buttons
    def end(self,obj):
        App.get_running_app().stop()
        Window.close()
        
    #Making a function to bind find lot
    def open_flot(self,obj):
        self.clear_widgets()
        self.add_widget(findlot())
        
    #Making a function to bind profile
    def open_profile(self,obj):
        self.clear_widgets()
        self.add_widget(profile())
    
    #This is a function for readding the orginal mm widgets
    def re_add(self):
        
        #Here we add a label for the app
        app_label = Label(text = "DigiPark", font_size = 75)
        self.add_widget(app_label)
        
        #Here we create 3 button widgets to the mm widget
        findlot = Button(text = "Find Lot",font_size = 75)
        profile = Button(text = "Profile",font_size = 75)
        exit1 = Button(text = "Exit", font_size = 75)
        
        #Adding the 3 Buttons to the main_menu widget
        self.add_widget(findlot)
        self.add_widget(profile)
        self.add_widget(exit1)
        
        
        #Binding the buttons to certain functions
        exit1.bind(on_press = self.end)
        findlot.bind(on_press = self.open_flot)
        profile.bind(on_press = self.open_profile)
        
        
        
    
        
        
    

#Widget/New Window for the Profile Window
class profile(GridLayout):
    
    #Initializing the widget object
    def __init__(self,**kwargs):
        
        #Calling the super classes constructor
        super(profile,self).__init__(**kwargs,rows = 4 , cols = 2, orientation ='tb-lr')
        
        #Here we add a image widget for profile picture
        profile_im = Image()
        self.add_widget(profile_im)
        
        #Here we add a email label widget for user's email
        email_l1 = Label(text = "email:",font_size = 75)
        self.add_widget(email_l1)
        
        #Here we add a username label widget for user's username
        user_l1 = Label(text = "username:",font_size = 75)
        self.add_widget(user_l1)
        
        #Here we add a single back button to widget
        back = Button(text = "back",font_size =75)
        self.add_widget(back)
        
        #Binding button to fuction for main menu
        back.bind(on_press = self.re_add_mm)
        
        #Here we add the profile label for the page
        p_label = Label(text = "Profile", font_size = 75)
        self.add_widget(p_label)
        
        #Here we add a label that will give the actual email
        email_l2 = Label(text = "", font_size = 40)
        self.add_widget(email_l2)
        
        #Here we add a label that will give the actual username
        user_l2 = Label(text = "", font_size = 40)
        self.add_widget(user_l2)
        
        #Here we add a image to display the last user locations
        image_loc = Image()
        self.add_widget(image_loc)
        
        
        #Thats all the widgets for the profile page
        
        
    
    #Here were going to bind the button to a function that gives back the main menu widgets
    def re_add_mm(self,obj):
        self.clear_widgets()
        self.parent.re_add()
        self.parent.remove_widget(self)
        
        


#Widget/New Window for the FindLot Window
class findlot(GridLayout):
    #Initializing the widget object
    def __init__(self,**kwargs):
        
        #Calling the super classes constructor
        super(findlot,self).__init__(**kwargs,rows = 5)
        
        
        #Here we add a Label for FindLot
        fl_label = Label(text = "Find Parking Lot", font_size = 75)
        self.add_widget(fl_label)
        
        #Here we add a search text input widget to allow user to search for parking lots
        search = TextInput(text = "search street address", font_size = 20)
        #Making search a instance variable so we can access it from open_pmap() version 3
        self.search = search
        self.add_widget(search)
        
        #4/11/2023 Adding a Gridlayout to put text inputs into
        self.scrollgrid = GridLayout(rows = 3,cols=1,orientation = 'tb-lr',size_hint_y = (None))
        self.scrollgrid.bind(minimum_height = self.scrollgrid.setter('height'))
        #4/11/2023 update to add scroll view and grod layout for scroll view
        self.scroll = ScrollView(size_hint = (1,None),size = (Window.width/2,Window.height/2))

        #Here we read the text file containing the addresses to display them to user, 4/11/2023 update to directory
        addr_list = open('address_list.txt','r')
        addrs = addr_list.readlines()
        for i in range(len(address_list)):
            ti = TextInput(text = (addrs[i]),size_hint_y = None, height = 60,readonly = True,allow_copy = True)
            self.scrollgrid.add_widget(ti)
        
        #Now adding the gridlayout to the scroll view so we can utilize it
        self.scroll.add_widget(self.scrollgrid)
        self.add_widget(self.scroll)
        
        #Here we add a single back button to widget
        back = Button(text = "back",font_size = 75)
        self.add_widget(back)
        
        #Here we add a button to go to parking lot map
        pmap = Button(text = "map",font_size = 75)
        self.add_widget(pmap)
        
        #Binding button to function for main menu
        back.bind(on_press = self.re_add_mm)
        pmap.bind(on_press = self.open_pmap)
        
        
    #Here were going to bind the button to a function that gives back the main menu widgets
    def re_add_mm(self,obj):
        self.clear_widgets()
        self.parent.re_add()
        self.parent.remove_widget(self)
        
    #Here we add a function to clear widget contents and open the map window
    #In version 3 we add reading the text input to see what adress user typed in
    def open_pmap(self,obj):
        
        #Here we fetch the text input widgets text
        addr = self.search.text
        
        self.clear_widgets()
        
        #Here we pass the addr to p_map to know what (address) user chose (version3)
        self.add_widget(p_map(addr))
        
    #Here we define a add function to readd the widgets of find_lot window
    def add(self):
        
        #Here we add a Label for FindLot
        fl_label = Label(text = "Find Parking Lot", font_size = 75)
        self.add_widget(fl_label)
        
        #Here we add a search text input widget to allow user to search for parking lots
        self.search = TextInput(text = "search street address", font_size = 20)
        self.add_widget(self.search)
        
        #4/11/2023 Adding a Gridlayout to put text inputs into
        self.scrollgrid = GridLayout(rows = 3,cols=1,orientation = 'tb-lr',size_hint_y = (None))
        self.scrollgrid.bind(minimum_height = self.scrollgrid.setter('height'))
        #4/11/2023 update to add scroll view and grod layout for scroll view
        self.scroll = ScrollView(size_hint = (1,None),size = (Window.width/2,Window.height/2))

        #Here we read the text file containing the addresses to display them to user, 4/11/2023 update to directory
        addr_list = open('address_list.txt','r')
        addrs = addr_list.readlines()
        for i in range(len(address_list)):
            ti = TextInput(text = (addrs[i]),size_hint_y = None, height = 40,readonly = True,allow_copy = True)
            self.scrollgrid.add_widget(ti)
        
        #Now adding the gridlayout to the scroll view so we can utilize it
        self.scroll.add_widget(self.scrollgrid)
        self.add_widget(self.scroll)
        
        #Here we add a single back button to widget
        back = Button(text = "back",font_size = 75)
        self.add_widget(back)
        
        #Here we add a button to go to parking lot map
        pmap = Button(text = "map",font_size = 75, size = (100,40))
        self.add_widget(pmap)
        
        #Binding button to fuction for main menu
        back.bind(on_press = self.re_add_mm)
        
        #Binding pmap to open_pmap()
        pmap.bind(on_press = self.open_pmap)
        
    
        
        
#Class for Parking Map Widget
class p_map(GridLayout):
    
    #Initializing the widget object
    def __init__(self,addr,**kwargs):
        
        #Calling the super classes constructor
        super(p_map,self).__init__(**kwargs,rows=5)
        
        #Creating a variable called count to count num time map is drawn
        self.count = 0
        
        #Here we get the user address (version 3 updates)
        self.addr = addr
        
        #Making a varible state to see if address is valid
        self.state = 0
        
        #Started 3/20/2023
        #1.) Format into request to the server "/spotarray/addr"
        req = "/spotarray/"+str(addr)
        
        #Fetching the appropriate ip address for the parking lot, 4/20/2023
        index = address_list.index(addr)
        ip = ip_addr_list[index]
        #All ports are 9999
        
        #2.) Socket connection to the server
        client = client_net(ip,9999)
        client.connect()
        
        #3.)Send the request to the server
        req = client.format_msg(req)
        client.send(req)
        
        #4.)Recieve the response to the server
        resp = client.recv()
    
        #5.) Extract the spotarray from the response "/spotarray/(spotarray int)"
        resp = client.unformat_msg(resp)
        spotarray = ""
        for i in range(len(resp) - 11):
            spotarray = spotarray + resp[i+11]
        #6.) If the address is invalid return too find lot window
        if(spotarray == "Invalid_Address"):
            print("issue")
        
        #7.) decimal to binary converter function
        spotarray = int(spotarray)
        spotarray = dec_to_bin(spotarray,29)
        #End of Client Functionality
        
              
        
        #Now based on the adress typed in we find a index
        try:
            print(self.addr)
            index = address_list.index(self.addr)
            #Fetch the image based off the index (3/20/2023)
            image = image_list[index]
            #Fetch the pos array for the image
            pos = pos_list[index]
            #Removing all previous image
            imlist = os.listdir()
            for i in range (len(address_list)):
                if(imlist.count(address_list[i] + str(self.count) + ".png") > 0):
                    os.remove(ref +"Drawn_Maps"+ address_list[i]+str(self.count) +".png")
            #Drawing the map and saving it
            im1 = draw_map(image,marker_list[0],marker_list[1],spotarray,pos)
            im1.save(ref +"Drawn_Maps" + str(addr) + str(self.count) + ".png")
            
        except(ValueError):
          #Just setting Image index to 0 by default if invalid address is typed
          print("Invalid Address, DigiPark does not cover this Lot")
          index = 0
          
        
        
        #Here we we add a image widget to display the actual parking lot map
        #Here the image will be our self.im we got from image_list (version 3 update) we use source = self.im.filename for image path
        im_source = ref +"Drawn_Maps" + str(addr) + str(self.count) + ".png"
        self.map_image = Image(source = im_source,allow_stretch = True)
        self.count = self.count +1
        self.add_widget(self.map_image)
        
        #4/9/2023 update adding gridlayout to for other widget other than image
        self.gl2 = GridLayout(rows = 6, cols =1, orientation = 'tb-lr')
        self.add_widget(self.gl2)
        
        #Here we add a Label for the parking map
        map_label = Label(text = "Parking Lot Map", font_size = 40)
        self.gl2.add_widget(map_label)
        
        #Here we add a 3 label to show parking statistics
        self.stat1_label = Label(text = "Num Spots: ", font_size = 40)
        self.gl2.add_widget(self.stat1_label)
        self.stat2_label = Label(text = "%Occupied: ",font_size = 40)
        self.gl2.add_widget(self.stat2_label)
        self.stat3_label = Label(text = "Num Open: ",font_size = 40)
        self.gl2.add_widget(self.stat3_label)
        
        #Computing statistics and formatting them into string, 4/9/2023 update
        num_spots = self.compute_num(spotarray)
        stat1 = "Num Spots: " + num_spots
        self.stat1_label.text = stat1
        per_occupied = self.compute_per_occupied(spotarray)
        stat2 = "%Occupied: "+ per_occupied
        self.stat2_label.text = stat2
        num_open = self.compute_num_open(spotarray)
        stat3 = "Num Open: " + num_open
        self.stat3_label.text = stat3
        
        
        #4/9/2023 update, here we add a refresh button for map window
        self.refresh = Button(text = "refresh", font_size = 40)
        self.gl2.add_widget(self.refresh)
        
        #Here we add a single back button to widget
        back = Button(text = "back",font_size = 40)
        self.gl2.add_widget(back)
        
        #Binding button to fuction for main menu
        back.bind(on_press = self.re_add_fl)
        #Now we bind the refresh button to the refresh function to refresh the map/statistics
        self.refresh.bind(on_press = self.refresh_funct)
    
    #Here were going to bind the button to a function that gives back the find lot widgets
    def re_add_fl(self,obj):
        self.clear_widgets()
        self.parent.add()
        self.parent.remove_widget(self)
        
    #For returning to the fl window when error occurs 
    def re_add_fl2(self):
        self.clear_widgets()
        self.parent.add()
        self.parent.remove_widget(self)
        
        
    
    #Here we define 3 functions to update the statistics of the lot based of spot array
    def compute_num(self,spotarray):
        return str(len(spotarray))
    
    def compute_per_occupied(self,spotarray):
        #computing num occupied
        sum = 0
        for i in range(len(spotarray)):
            if(spotarray[i] == 1):
                sum = sum +1
            else:
                sum = sum
        #Computing percentage
        sum = (sum/len(spotarray))*100
        return str(sum)
    
    def compute_num_open(self,spotarray):
        #computing num unoccupied
        sum = 0
        for i in range(len(spotarray)):
            if (spotarray[i] == 0):
                sum = sum +1
            else:
                sum = sum
        return str(sum)
    
    #4/9/2023 Update to Parking app, Adding refresh function to refresh the page
    def refresh_funct(self,spotarray):
        
        #Refetching the spot array form server for parkinglot_address
         #1.) Format into request to the server "/spotarray/addr"
        req = "/spotarray/"+str(self.addr)
        
        #Fetching the appropriate ip address for the server, 4/20/2023
        index = address_list.index(self.addr)
        ip = ip_addr_list[index]
        
        #2.) Socket connection to the server
        client = client_net(ip,9999)
        client.connect()
        #3.)Send the request to the server
        req = client.format_msg(req)
        client.send(req)
        #4.)Recieve the response to the server
        resp = client.recv()
        #5.) Extract the spotarray from the response "/spotarray/(spotarray int)"
        resp = client.unformat_msg(resp)
        spotarray = ""
        for i in range(len(resp) - 11):
            spotarray = spotarray + resp[i+11]
        print(spotarray)
        spotarray = int(spotarray)
        #6.) decimal to binary converter function
        spotarray = dec_to_bin(spotarray,29)
       
       
       #Now redrawing the parking map image with new spotarray and updating the image widget
        try:
            index = address_list.index(self.addr)
            #Fetch the image based off the index (3/20/2023)
            image = image_list[index]
            #Fetch the pos array for the image
            pos = pos_list[index]
            #Removing all previous image
            imlist = os.listdir()
            for i in range (len(address_list)):
                if(imlist.count(address_list[i] + str(self.count) + ".png") > 0):
                    os.remove(ref +"Drawn_Maps" + address_list[i]+ str(self.count) + ".png")
            #Drawing the map and saving it
            im1 = draw_map(image,marker_list[0],marker_list[1],spotarray,pos)
            im1.save(ref + "Drawn_Maps" + str(self.addr) + str(self.count) +".png")
            
        except(ValueError):
          #Just setting Image index to 0 by default if invalid address is typed
          print("Invalid Address, DigiPark does not cover this Lot")
          index = 0
        #Updating the parking map image
        im_source = ref +"Drawn_Maps"+ str(self.addr) + str(self.count) + ".png"
        self.map_image.source = im_source
        self.count = self.count+1
        
        #Now we update the parking lot statistics with new spotarray
        #Computing statistics and formatting them into string
        num_spots = self.compute_num(spotarray)
        stat1 = "Num Spots: " + num_spots
        self.stat1_label.text = stat1
        per_occupied = self.compute_per_occupied(spotarray)
        stat2 = "%Occupied: "+ per_occupied
        self.stat2_label.text = stat2
        num_open = self.compute_num_open(spotarray)
        stat3 = "Num Open: " + num_open
        self.stat3_label.text = stat3
          
          
           
    
        
        
    
####################################################################################################################################    
#Here we define a unique class called (client_net), that will handle all the client/server functionality 3/20/2023
class client_net():
    
    #Initialization of the class
    def __init__(self,host,port):
        
        #Setting the host and the port of the client
        self.host = host
        self.port = port
        self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    
    #Here we define a function for making a socket connection to the host/port
    def connect(self):
        #Connecting to the server
        self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        self.sock.connect((self.host,self.port))
        
    #Here we define a function that will take in string and format it to be sent, 64 bits
    def format_msg(self,msg):
        #formatting the msg to utf-8 and padding it
        msg = msg.encode('utf-8')
        msg = msg + b''*(120 - len(msg))
        return msg
    
    #Here we define a function that will send the formatted message to the server
    def send(self,msg):
        self.sock.sendall(msg)
    
    #Here we define a function that will recv a message from the server
    def recv(self):
        msg = self.sock.recv(120)
        return msg
    
    #Here we define a function that will decode and return a received message
    def unformat_msg(self,msg):
        msg = msg.decode('utf-8')
        return msg
#The functions above will do all the basic (server/client)  functionality






#Convert decimal values to binary, 3/20/2023
def dec_to_bin(dec, size):
    #size = 8 #temp size for parking lot
    bin_array = [0 for x in range(size)]
    i = size-1
    
    while(dec > 0):
        bin_array[i] = dec % 2
        
        dec = int(dec / 2)
        i -= 1
    
    #print(bin_array)
    
    return bin_array

def check_addr(addr):
    i = 0
    size = len(address_list)
    
    while(i < size):
        if address_list[i] == addr:
            return i
        
        i += 1

    return size-1
###################################################################################################################
    
       
        

#Now for the parking app to start running we check for main function
if __name__ == "__main__":
    #Create a new park_app object and run it
    p_app1 = park_app()
    p_app1.run()
        
    
        



