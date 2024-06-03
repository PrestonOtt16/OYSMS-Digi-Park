#This is integration of the Parking_Data_Gatherer and the ML_Model, this program will periodically capture images
#, segement them using a mask, run the ML Model on each segement, and draw the predictions on the original image
#Preston Ott, OYSMS Digi-Park



#Importing the libraries we'll need
import os
from PIL import Image as im
import time
import openpyxl as xl
from picamera2 import Picamera2, Preview
import numpy
import threading
import pickle
import cv2
import pandas
import sklearn



#Importing the libraries we'll need to make the app
import kivy
import sqlite3
from PIL import Image as im
from PIL import ImageDraw
import os
import math
from kivy.core.window import Window
from kivy.app import App
from kivy.uix.widget import Widget
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.core.window import Window
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.relativelayout import RelativeLayout
from kivy.uix.image import Image
from kivy.uix.textinput import TextInput
from kivy.uix.textinput import TextInput
from kivy.clock import Clock

################################################################################################################
#Here we define function fetch_mask to fetch mask from a csv file for our particular parking lot
def fetch_mask(path,numspots):
    
    #Fetching the specific csv file
    book1 = xl.load_workbook(path)
    
    #Now we get the active worksheets in this workbook
    sheet1 = book1.active
    
    #Now we get make a numpy array to store the mask
    mask_lot = numpy.zeros((numspots,4))
    
    #Now we store the mask in the numpy array
    for i in range(numspots):
        mask_lot[i,:] = [sheet1.cell(i+1,1).value, sheet1.cell(i+1,2).value, sheet1.cell(i+1,3).value, sheet1.cell(i+1,4).value]
    
    #Now we return the mask we made
    return mask_lot
#Here we define a function that will segment a image into individual spots using the mask
def segment_image(mask,image,image_path):
    
    #First we create an array to store the images
    image_array = []
    os.chdir(image_path)
    park_lot_image = image
    
    
    #Now we segment the (image) into individual spots and scale them to (30x40) pixels
    print(mask.shape[0])
    for i in range(mask.shape[0]):
        #Extracting the ith image
        image_array.append(park_lot_image.crop(box = (mask[i,:])))
        
    #Now we resize the images to be (30x40) pixels
    for j in range(len(image_array)):
        #Resizing the image to be (30x40) pixels
        image_array[j] = image_array[j].resize(size = (30,40))
        
    #Returning the list of images
    return image_array
    
#Here is a function that will take in a image object and apply the mask to it
def mask_image(mask,image):
    
    #Defining a image to draw on
    im2 = ImageDraw.Draw(image)
    
    #First we go into a loop size of mask
    for i in range(mask.shape[0]):
        rect = tuple(mask[i,:])
        im2.rectangle(rect,fill = None,outline = (0,255,0))
    
    #Now we return the image object
    return image
        
    
        

#Now we make a function that will save a list of images to the unlabeled data folder
def save_images(path,image_list):
    
    #First we change our cwd to path
    os.chdir(path)
    
    #Now we get the list of items in folder
    list = os.listdir()
    
    #Now we save all the images and give them ordered names
    for i in range(len(image_list)):
        #Making name for image
        name = "image" + str(len(list)+i) + ".jpg"
        image_list[i].save(name)
    #Now we've sucessfully updated the unlabeled data
        
#Here we define a function for the RPI4 to capture a image with the OV5647
def capture_image(path,imstate):
    
    #First setting our directory to path
    os.chdir(path)
    
    #Creating a picamera object and initializing it
    picam2 = Picamera2()
    preview_config = picam2.create_preview_configuration(main={"size": (1280, 720)})
    picam2.configure(preview_config)
    picam2.start()
    time.sleep(1)
    metadata = picam2.capture_file("im_lot"+str(imstate)+".jpg")
    park_lot_image = im.open("im_lot"+str(imstate)+".jpg")
    picam2.close()
    return park_lot_image
    
#Here we define a function that will check if a folder(path) exceeds maximum memory
def max_memory(max_bytes,path):
    
    #First we change directories to the path
    os.chdir(path)
    #Now we fetch the size of the folder connected to the path
    num_bytes = os.path.getsize(path)
    
    #Now we check if num_bytes exceeds max_bytes
    if(num_bytes >= max_bytes):
        return 1
    else:
        return 0
##############################################################################################################
#Now we import the ML Model and the functions we'll need to make predictions

#Here we load in the ML Model
file1 = open('ML_Model_V1.2.obj','rb')
ann1 = pickle.load(file1)

#Here we utitlize a built in previous function from Park_IM_Pro.py that is modified to convert the images into 1d rgb arrays
def im_array1d(im):
    #value_index  = 120*i+3*j+k
    im_array = numpy.zeros(3600)
    for i in range(30):
        for j in range(40):
            for k in range(3):
                #Writing a RGB value into the array
                im_array[120*i +3*j +k] = im[j-1,i-1,k-1]        
    return im_array



#Here we make a function that will take in a image and resize them to 30x40 images, then convert each image into a 3d rgb array
def process_images(image,mask):
    
    #Here we segement the image with the mask and resize each image to be (30x40) in size
    image_list = []
    for i in range(mask.shape[0]):
        image_list.append(image.crop(box = mask[i,:]))
        image_list[i] = image_list[i].resize(size = (30,40))
        #Here we save the images in the ML_Model_Segements
        os.chdir("Machine-Learning-Model-Application/Data")
        image_list[i].save("image" + str(i) + ".jpg")
    
    #Now we take each image and convert it into a 3d numpy array
    image_rgb_list = []
    for i in range(len(image_list)):
         image_rgb_list.append(cv2.imread(image_list[i]))
         #Now we delete the images in the ML_Model_Segments folder
         os.chdir("Machine-Learning-Model-Application/Data")
         os.remove("image" + str(i) + ".jpg")
         
    #Now we take each 3d array and convert into a 1d array [r1,g1,b1,r2,g2 ... b1200] for each image
    image_rgb1d_list = []
    for i in range(len(image_list)):
        image_rgb1d_list.append(im_array1d(image_rgb_list[i]))
        
    #Now we store this list of 1d arrays in a pandas dataframe, each 1d list corresponds to an image
    image_df = pandas.DataFrame(image_rgb1d_list)
    return image_df


#Here we make a function that will take in the pandas dataframe and utilize the ML Model to produce a spot_array
def predict_spotarray(dataframe,model):
    
    #Here we a list for the spot array
    spot_array = []
    #Now we go through each row in pandas dataframe and make a prediction
    for i in range(dataframe.shape[0]):
        spot_array.append(model.predict(dataframe.iloc[[i]]))
    #Now we return the spot array
    return spot_array


#Here we make a function that will take in a image and draw the mask according to the spotarray (green,red) = (unoccupied/occupied)
def draw_mask(image,spotarray,mask):
    
    #Here define a Draw object for the image to draw mask
    im2 = ImageDraw.Draw(image)
    
    #Now we go into a for loop to draw each rectangle on the mask for our image
    for i in range(mask.shape[0]):
        
        #Getting the ith rectangle for the mask
        rect = mask[i,:]
        
        #Now we determine if rectangle is red or green
        if(spotarray[i] == 0):
            im2.rectangle(rect,fill = None,outline = (0,255,0))
        else:
            im2.rectangle(rect,fill = None,outline = (255,0,0))
        
    #Now we return the image that now has the mask drawn upon it
    return image
            
        
        
    
##############################################################################################################
#Now we define a kivy app with a kivy widget
    
#Here we define a class which is a child of the kivy app class that will handle generating our GUI
class data_gatherer_app(App):
    #building app, by returning/generating root widget
    def build(self): 
        #making/returning root widget
        mm = main_menu()
        return mm
    
#Now we define a kivy widget to be our root widget
class main_menu(GridLayout):
    
    #Here is the constructore for the main menu widget
    def __init__(self,**kwargs):
        
        #calling super classes constructor, we make the float layout the size of the window
        super(main_menu,self).__init__(**kwargs,rows = 2,orientation = 'tb-lr',size = Window.size)
        
        #Theser are instance variables that hold the user parameters
        self.csv_path = ""
        self.data_path = ""
        self.image_path = ""
        self.num_spots = 0
        self.time = 1
        self.impath = ""
        #This variable will control if the app continues to capture data
        self.state = 1
        #Here we make a variable to alternate the path of the image taken
        self.imstate = 1
        #Here we add a ML Model as a parameter
        self.model = ann1
        
        #Now we add a image widget and gridlayout widget for level 1 in widget tree
        self.park_lot_image = Image(keep_ratio = False)
        self.add_widget(self.park_lot_image)
        gl1 = GridLayout(cols = 2,orientation = 'tb-lr')
        self.add_widget(gl1)
        
        #Now we add 2 grid layouts to gridlayot 1 to organize widgets in level2 of widget tree
        gl2 = GridLayout(rows = 4,orientation = 'tb-lr')
        gl1.add_widget(gl2)
        gl3 = GridLayout(cols = 2,orientation = 'tb-lr')
        gl1.add_widget(gl3)
        
        #Now we add 1 label and the 3 buttons for gl2 and 2 gridlayouts for gl3 on layer 3 of widget tree
        t_label = Label(text = "Digi-Park Data-GatherApp",font_size = 20)
        gl2.add_widget(t_label)
        start = Button(text = "start")
        gl2.add_widget(start)
        stop = Button(text = "stop")
        gl2.add_widget(stop)
        exit_ = Button(text = "exit")
        gl2.add_widget(exit_)
        
        gl4 = GridLayout(rows = 5,orientation = 'tb-lr')
        gl3.add_widget(gl4)
        gl5 = GridLayout(rows = 5, orientation = 'tb-lr')
        gl3.add_widget(gl5)
        
        
        #Now we add the labels for the text inputs and the text inputs themselves on level 4 of widget tree
        mask_label = Label(text = "mask_csv_file_path: ",font_size = 15)
        gl4.add_widget(mask_label)
        data_label = Label(text = "unlabeled_datafolder_path: ",font_size = 15)
        gl4.add_widget(data_label)
        image_label = Label(text = "parking_image_path: ",font_size = 15)
        gl4.add_widget(image_label)
        numspots_label = Label(text = "num_spots: ",font_size = 15)
        gl4.add_widget(numspots_label)
        time_label = Label(text = "sampling_time[s]: ",font_size = 15)
        gl4.add_widget(time_label)
        
        self.mask_text = TextInput(font_size = 10, text = "1")
        gl5.add_widget(self.mask_text)
        self.data_text = TextInput(font_size = 10,text = "1")
        gl5.add_widget(self.data_text)
        self.image_text = TextInput(font_size = 10,text = "1" )
        gl5.add_widget(self.image_text)
        self.numspots_text = TextInput(font_size = 10)
        gl5.add_widget(self.numspots_text)
        self.time_text = TextInput(font_size = 10)
        gl5.add_widget(self.time_text)
        
        #Here we make start_capture a event that will be called every (1/20s)
        event = Clock.schedule_interval(self.start_capture,2)
        
        
        #Now we bind functions to our 3 buttons
        exit_.bind(on_press = self.close_app)
        start.bind(on_press = self.start)
        stop.bind( on_press = self.stop_capture)
        
        
    #This is a basic function to close the app
    def close_app(self,obj):
        #Get the current running app(stop it) and close window
        App.get_running_app().stop()
        Window.close()
        
    #Here we define a function for updating the state variable with start button
    def start(self,obj):
        self.state = 0
        
        
            
    #Here we define a function that will capture/process/save images of the parking lot
    #This function will be a callback and will be called every (1/20)s to get more image data
    def start_capture(self,obj):
       
        #Here we add a conditional statment to check if user has started capture
        if(self.state == 0):
            #Fetching the text input contents
            self.csv_path = self.mask_text.text
            self.data_path = self.data_text.text
            self.image_path = self.image_text.text
            self.num_spots = int(self.numspots_text.text)
            self.time = int(self.time_text.text)
                
            #We fetch the mask from the csv file
            mask = fetch_mask(self.csv_path,self.num_spots)
            #Now we capture ans image and save it in ipath
            im1 = capture_image(self.image_path,self.imstate)
            #Now we segment the image into indiviudal spots
            im_list = segment_image(mask,im1,self.data_path)
            #Now we save the images into the unlabeled data folder
            save_images(self.data_path,im_list)
            #Now we check if the folder in data_path exceeds maximum storage 1GB
            state = max_memory(1000000,self.data_path)
            print(state)
            if(state != 0):
                #If state is not 0 we exceeded max memory closing app
                App.get_running_app().stop()
                Window.close()
            
            #Now we start utilizing the ML Model,first we process the image taken into a pandas dataframe of images
            dataframe = process_images(im1,mask)
            #Now we make predictions on each image in the dataframe to get the spot array
            spotarray = predict_spotarray(dataframe,self.model)
            #Now we draw the mask on the image according to the spotarray
            masked_image = draw_mask(im1,spotarray,mask)
            
            #Now we delete the current image in the image folder and put in our masked image
            os.chdir("/home/user1/Desktop/Data_Gatherer/Image_Data")
            if(self.imstate != 1):
                os.remove(str(self.image_path)+"/im_lot"+str(self.imstate-1)+".jpg")
            masked_image.save(str(self.image_path)+"/im_lot"+str(self.imstate)+".jpg")
            #Updating the image state for the next image we take and the image widget source image
            self.park_lot_image.source = (str(self.image_path)"+/im_lot"+str(self.imstate)+".jpg")
            self.imstate = self.imstate +1
            
                
            
        #If user wants to stop capture
        else:
            print("capturing stopped")
            
    #Here we define a function that will stop the capture/process/save function
    def stop_capture(self,obj):
        #Here we simply update the self.state to 1 to stop capturing
        print(self.state)
        self.state = 1
    
    
        
    
        

#Here is the main function of the program that will run in a infinite loop
if __name__ == "__main__":
    
    #Create a new park_app object and run it
    l_data = data_gatherer_app()
    #l_data.run()
    thread1 = threading.Thread(target = l_data.run())
    thread1.start()
    
